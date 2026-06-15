"""Ingest Ravensburger Lorcana tournament events into SQLite.

Usage:
    python ingest.py --xlsx "C:\\path\\to\\events.xlsx" --season "Fabled Fall 2025"
    python ingest.py --ids 198237 203889 ...
    python ingest.py --urls https://tcg.ravensburgerplay.com/events/198237 ...

Idempotent: re-running skips events that already have matches recorded.
"""
import argparse, json, re, sqlite3, sys, time, urllib.request, urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

HERE = Path(__file__).parent
DB_PATH = HERE / "lorcana_elo.db"
SCHEMA = HERE / "schema.sql"
API = "https://api.cloudflare.ravensburgerplay.com/hydraproxy/api/v2/player/events/{eid}"
API_META = "https://api.ravensburgerplay.com/api/v2/events/{eid}/"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}


def http_get(url, retries=3):
    last = None
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.load(r)
        except Exception as e:
            last = e
            time.sleep(0.5 * (i + 1))
    raise last


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = db()
    conn.executescript(SCHEMA.read_text())
    conn.commit()
    conn.close()


def get_or_create_player(conn, display_name, first_event_id):
    row = conn.execute(
        "SELECT player_id FROM players WHERE display_name = ?", (display_name,)
    ).fetchone()
    if row:
        return row["player_id"]
    cur = conn.execute(
        "INSERT INTO players (display_name, first_event_id) VALUES (?, ?)",
        (display_name, first_event_id),
    )
    return cur.lastrowid


def event_already_ingested(conn, event_id):
    """Skip an event only if it's FINISHED on the platform AND we already have
    matches for it. In-progress / not-started events keep getting re-pulled on
    weekly refreshes so newly entered rounds land in our DB."""
    row = conn.execute(
        """SELECT e.status, EXISTS(SELECT 1 FROM matches m WHERE m.event_id=e.event_id) AS has_matches
           FROM events e WHERE e.event_id = ?""", (event_id,),
    ).fetchone()
    if not row:
        return False
    return row["status"] == "EVENT_FINISHED" and bool(row["has_matches"])


def ingest_event(event_id, store=None, location=None, event_date=None, season=None):
    """Pull one event + all its rounds and write to DB. Idempotent at event level."""
    conn = db()
    try:
        # Honor manual ignore flag — don't re-pull events the user marked as did-not-run
        row = conn.execute("SELECT is_ignored FROM events WHERE event_id = ?", (event_id,)).fetchone()
        if row and row["is_ignored"]:
            return event_id, "skip", "ignored (marked did-not-run)"
        if event_already_ingested(conn, event_id):
            return event_id, "skip", "already ingested"

        tv = http_get(API.format(eid=event_id) + "/tv/")
        name = tv["name"]
        status = tv.get("lifecycle_status")
        num_players = tv.get("starting_player_count")
        phases = tv.get("tournament_phases", [])

        # When date wasn't provided (--ids mode), fetch it from the main events API.
        if event_date is None:
            try:
                meta = http_get(API_META.format(eid=event_id))
                raw_dt = meta.get("start_datetime") or ""
                if raw_dt:
                    event_date = raw_dt[:10]  # "2026-06-13T15:00:00+00:00" → "2026-06-13"
            except Exception:
                pass

        # collect all rounds with their type label + parent phase type (SWISS vs RANKED_SINGLE_ELIMINATION)
        all_rounds = []  # (round_id, round_number, round_type, phase_type, round_status)
        for ph in phases:
            ph_type = ph.get("round_type")  # API stores phase 'type' under round_type at phase level
            for r in ph.get("rounds", []):
                all_rounds.append((r["id"], r["round_number"], r.get("round_type"), ph_type, r.get("status")))

        # If the event hasn't run yet (or has no exposed rounds), still record
        # the event metadata so future refreshes know to come back. The first
        # time we ingest a not-yet-played event we want it queued, not skipped.
        if not all_rounds:
            with conn:
                conn.execute(
                    """INSERT OR REPLACE INTO events
                       (event_id, name, store, location, event_date, season, num_players, status)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (event_id, name, store, location, event_date, season, num_players, status),
                )
            return event_id, "queue", f"{name} (no rounds yet — queued for refresh)"

        # pull all rounds. An UPCOMING round that hasn't started yet (e.g. the
        # finals of a top cut while the semis are still running) has no matches
        # endpoint — RPH 404s it. Skip that one round instead of letting it abort
        # the whole event; it fills in on a later refresh once the round starts.
        # (Without this, any event mid-top-cut fails to ingest entirely.)
        round_payloads = {}
        for rid, _, _, _, _ in all_rounds:
            url = API.format(eid=event_id) + f"/tv/matches/?round_id={rid}"
            try:
                round_payloads[rid] = http_get(url)
            except urllib.error.HTTPError as e:
                if e.code == 404:
                    round_payloads[rid] = {"results": []}
                else:
                    raise

        # write everything in one transaction
        with conn:
            conn.execute(
                """INSERT OR REPLACE INTO events
                   (event_id, name, store, location, event_date, season, num_players, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (event_id, name, store, location, event_date, season, num_players, status),
            )

            # Dedup guard against a player's display-name spelling changing between
            # re-ingests of a not-yet-finished event. A changed name -> new
            # player_id, so the UNIQUE(...,player1_id) key no longer collides and
            # INSERT OR IGNORE would add a SECOND row for the same physical match
            # (there is exactly one table per round). Skip any (round, table) we
            # already have. Bit "LoL_METALLICFLARE" vs "LoL METALLICFLARE" once —
            # double-counted whole rounds across two pulls.
            existing_rt = {(r["round_id"], r["table_number"]) for r in conn.execute(
                "SELECT round_id, table_number FROM matches "
                "WHERE event_id=? AND is_bye=0 AND table_number IS NOT NULL", (event_id,))}

            for rid, rnum, rtype, phase_type, rstatus in all_rounds:
                d = round_payloads[rid]
                for m in d.get("results", []):
                    players = m.get("players", [])
                    is_bye = bool(m.get("match_is_bye"))

                    if is_bye:
                        if not players:
                            continue
                        p = players[0]
                        pid = get_or_create_player(conn, p["tv_display_name"], event_id)
                        conn.execute(
                            """INSERT OR IGNORE INTO matches
                               (event_id, round_id, round_number, round_type, table_number,
                                player1_id, player2_id, winner_id, games_won_p1, games_won_p2, is_bye)
                               VALUES (?, ?, ?, ?, ?, ?, NULL, ?, ?, NULL, 1)""",
                            (event_id, rid, rnum, rtype, m.get("table_number"),
                             pid, pid, p.get("games_won")),
                        )
                        continue

                    if len(players) < 2:
                        continue
                    tnum = m.get("table_number")
                    if tnum is not None and (rid, tnum) in existing_rt:
                        continue  # same physical match already recorded (see guard above)
                    p1, p2 = sorted(players, key=lambda p: p.get("player_order") or 0)
                    p1_id = get_or_create_player(conn, p1["tv_display_name"], event_id)
                    p2_id = get_or_create_player(conn, p2["tv_display_name"], event_id)
                    winner_id = None
                    if p1.get("is_winner"):
                        winner_id = p1_id
                    elif p2.get("is_winner"):
                        winner_id = p2_id
                    conn.execute(
                        """INSERT OR IGNORE INTO matches
                           (event_id, round_id, round_number, round_type, table_number,
                            player1_id, player2_id, winner_id, games_won_p1, games_won_p2, is_bye)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)""",
                        (event_id, rid, rnum, rtype, tnum,
                         p1_id, p2_id, winner_id,
                         p1.get("games_won"), p2.get("games_won")),
                    )

        # Record gaps: COMPLETE rounds with 0 exposed matches.
        # Even the API marks them complete, but no pairings were entered.
        gaps = 0
        with conn:
            for rid, rnum, rtype, phase_type, rstatus in all_rounds:
                n_in_db = conn.execute(
                    "SELECT COUNT(*) FROM matches WHERE event_id = ? AND round_id = ?",
                    (event_id, rid),
                ).fetchone()[0]
                if n_in_db == 0 and rstatus == "COMPLETE":
                    conn.execute(
                        """INSERT OR REPLACE INTO gap_rounds
                           (event_id, round_id, round_number, phase_type, expected_matches, filled_matches)
                           VALUES (?, ?, ?, ?, NULL, 0)""",
                        (event_id, rid, rnum, phase_type),
                    )
                    gaps += 1

        return event_id, "ok", f"{name} ({len(all_rounds)} rounds, {gaps} gaps)"
    except Exception as e:
        return event_id, "err", repr(e)
    finally:
        conn.close()


def load_from_xlsx(path):
    """Yield (event_id, store, location, event_date) tuples from the user's spreadsheet.

    Auto-detects the data sheet by looking for a row whose column A says 'Date'
    and column G says 'Play Hub Link' (case-insensitive). The user's per-season
    spreadsheets use either 'Events' or 'Sheet1' as the sheet name and sometimes
    reorder middle columns — but col A (date), C (store), D (location), G (link)
    have been stable across files."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # Map known header tokens to canonical column names. We accept several
    # variants because the spreadsheet maintainers reorganize columns between
    # seasons (e.g. ROJ had a "Location" column; Wilds Unknown split it into
    # City + State; Whispers used "Play Hub Link" while Wilds Unknown uses
    # "Registration Link"). Auto-detecting the layout instead of hardcoding
    # column positions means new season files just work.
    def find_layout(ws):
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            cols = {}
            for j, cell in enumerate(row):
                if not isinstance(cell, str): continue
                h = cell.strip().lower()
                if h.startswith("date"):                                  cols["date"]  = j
                elif h in ("store", "store name"):                        cols["store"] = j
                elif h == "city":                                         cols["city"]  = j
                elif h == "state":                                        cols["state"] = j
                elif h == "location":                                     cols["location"] = j
                elif h in ("play hub link", "registration link", "event link"):
                    cols["link"] = j
            if "date" in cols and "link" in cols:
                return i, cols
        return None, {}

    ws = None; hdr = None; cols = {}
    for candidate in wb.worksheets:
        h, c = find_layout(candidate)
        if h is not None:
            ws = candidate; hdr = h; cols = c; break
    if ws is None:
        raise ValueError(f"no data sheet found in {path} (need a header row with Date + Play Hub Link / Registration Link)")

    for r in ws.iter_rows(min_row=hdr + 1, values_only=False):
        if cols["link"] >= len(r):
            continue
        g = r[cols["link"]]
        link = (g.hyperlink.target if g.hyperlink else None) or g.value
        if not isinstance(link, str):
            continue
        m = re.search(r"/events/(\d+)", link)
        if not m:
            continue
        eid = int(m.group(1))
        date_val = r[cols["date"]].value if cols.get("date") is not None else None
        date_iso = date_val.date().isoformat() if hasattr(date_val, "date") else None
        store = r[cols["store"]].value if cols.get("store") is not None else None
        # Build location from City + State (new format) or the single Location cell (old).
        if "city" in cols:
            city = r[cols["city"]].value if cols["city"] < len(r) else None
            state = r[cols["state"]].value if "state" in cols and cols["state"] < len(r) else None
            loc = f"{city}, {state}" if (city and state) else (city or state or None)
        elif "location" in cols and cols["location"] < len(r):
            loc = r[cols["location"]].value
        else:
            loc = None
        yield eid, store, loc, date_iso


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", help="path to spreadsheet (uses col G for event URLs)")
    ap.add_argument("--ids", nargs="*", type=int, default=[])
    ap.add_argument("--urls", nargs="*", default=[])
    ap.add_argument("--season", default=None, help="season label to tag events with")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    init_db()
    work = []  # list of (eid, store, loc, date)
    if args.xlsx:
        work.extend(load_from_xlsx(args.xlsx))
    for eid in args.ids:
        work.append((eid, None, None, None))
    for u in args.urls:
        m = re.search(r"/events/(\d+)", u)
        if m:
            work.append((int(m.group(1)), None, None, None))

    if not work:
        print("no events to ingest. pass --xlsx, --ids, or --urls.", file=sys.stderr)
        sys.exit(1)

    print(f"ingesting {len(work)} events with {args.workers} workers...")
    counts = {"ok": 0, "skip": 0, "err": 0, "queue": 0}
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {
            ex.submit(ingest_event, eid, store, loc, date, args.season): eid
            for eid, store, loc, date in work
        }
        for f in as_completed(futs):
            eid, status, msg = f.result()
            tag = {"ok":"OK  ","skip":"SKIP","err":"ERR ","queue":"QUEUE"}[status]
            print(f"  {tag} {eid}  {msg}")
            counts[status] += 1

    print(f"\nDone. ok={counts['ok']} queue={counts['queue']} skip={counts['skip']} err={counts['err']}")


if __name__ == "__main__":
    main()
